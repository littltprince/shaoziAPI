#encoding utf-8
from openpyxl import load_workbook
from method.ReadConfig import opconfig
import os
from method.Get_message import Get_message
base_dir=os.path.dirname(os.path.dirname(__file__))
data_path=os.path.join(base_dir,'data/ddt_data.xlsx')
filePath=os.path.join(base_dir,'configs/test_case.config')
class operationpyxl:
    def __init__(self,path,sheet_name):
        self.path=path
        self.sheet_name=sheet_name
    def get_data(self):
        a=load_workbook(self.path)
        sheet=a[self.sheet_name]
        # print(sheet.max_row)
        # print(sheet.max_column)
        '''1、根据用例编号来执行用例'''
        testnum=opconfig().OpConfig(filePath,'PEOPLE','testnum')
        '''2、根据模块名来执行用例'''
        mode=opconfig().OpConfig(filePath,'PEOPLE','mode')
        '''切换测试的接口环境
            Api----》正式环境
            TestApi---->测试环境'''
        host=getattr(Get_message,'Api')
        test_data=[]
        for i in range(2,sheet.max_row+1):#从第2行开始取值
            sub_data = {}
            sub_data["case_id"]=sheet.cell(i,1).value
            sub_data["mode"]= sheet.cell(i,2).value
            sub_data["title"] = sheet.cell(i,3).value
            # sub_data["url"] = sheet.cell(i,4).value
            if sheet.cell(i,4).value.find('${host}') !=-1:
                sub_data["url"]=sheet.cell(i,4).value.replace('${host}',host)
            else:
                sub_data["url"]= sheet.cell(i,4).value
            sub_data["data"] = sheet.cell(i,5).value
            sub_data["method"] = sheet.cell(i,6).value
            sub_data["expect"] = sheet.cell(i,7).value
            sub_data["fact"] = sheet.cell(i,8).value
            test_data.append(sub_data)
        '''1、根据用例编号来执行用例'''
        if testnum =='all':
            final_data=test_data
        else:
            final_data=[]
            for item in test_data:
                if item['case_id'] in eval(testnum):
                    final_data.append(item)
        return final_data
# if __name__ == '__main__':
#     print(operationpyxl(data_path,'login').get_data())
#     print(operationpyxl(data_path,'my').get_data())
        # '''2、根据模块来执行用例'''
        # if mode == None:
        #     final_data = test_data
        # else:
        #     final_data = []
        #     for item in test_data:
        #         if item['mode'] in mode:
        #             final_data.append(item)
        # return final_data
    '''写会用例的实际结果'''
    def write_fact(self,row,fact):
        a=load_workbook(self.path)
        sheet=a[self.sheet_name]
        sheet.cell(row,8).value=fact
        a.save(self.path)
    '''写回用例的执行通过或者失败,excel中只能写入str和数字类型的数据，其他数据的需要转格式'''
    def write_result(self,row,result):
        a=load_workbook(self.path)
        sheet=a[self.sheet_name]
        sheet.cell(row,9).value=result
        a.save(self.path)





# if __name__ == '__main__':
#     print(operationpyxl(data_path,'sheet1').get_data())


