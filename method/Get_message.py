#encoding:utf-8
from openpyxl import load_workbook
import os
base_dir=os.path.dirname(os.path.dirname(__file__))
data_path=os.path.join(base_dir,'data/ddt_data.xlsx')

class Get_message:
    Cookie=None
    Api=load_workbook(data_path)['api'].cell(1,2).value
    TsestApi=load_workbook(data_path)['api'].cell(2,2).value

